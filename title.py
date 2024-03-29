import rarfile
import pytesseract
from PIL import Image
import os
import pandas as pd
from openpyxl import load_workbook

rarfile.UNRAR_TOOL = 'C:\\Program Files\\WinRAR\\unrar.exe'
pytesseract.pytesseract.tesseract_cmd = r'C:\Program Files\Tesseract-OCR\tesseract.exe'

def extract_title_from_text(text):
    lines = text.strip().split('\n')
    if lines:
        return lines[0]
    else:
        return ""

def extract_text_from_image(image_file):
    try:
        with Image.open(image_file) as img:
            text = pytesseract.image_to_string(img)
            return extract_title_from_text(text)
    except PermissionError as e:
        print(f"PermissionError: {e}. Skipping file: {image_file}")
        return ""

def extract_titles_from_images(image_files):
    titles = []
    for file in image_files:
        title = extract_text_from_image(file)
        titles.append(title.strip())
    return titles

def extract(rar_file_path, folder_name):
    with rarfile.RarFile(rar_file_path, 'r') as rf:
        image_files = [rf.extract(file, path='temp') for file in rf.infolist() if os.path.dirname(file.filename) == folder_name]
        titles = extract_titles_from_images(image_files)
    return titles

rar_file_path = 'scientific_publication.rar'
folder_name = 'scientific_publication'  # Specify the folder path relative to the RAR file
titles = extract(rar_file_path, folder_name)

excel_file_path = 'Output_format.xlsx'
wb = load_workbook(excel_file_path)
ws = wb.active


column = 'Document Title'


for i, author in enumerate(authors, start=1):
    ws[f'{column}{i}'] = author


wb.save(excel_file_path)

print(f"Author names saved to {excel_file_path}")

title
